from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors

def handelExcel():
    # excel_address = r"cases.xlsx"
    excel_address = r"C:\Users\Administrator\Desktop\cases.xlsx"
    wb = load_workbook(excel_address)
    sht = wb.worksheets[0]

    fill_1 = PatternFill("solid", fgColor=colors.YELLOW)

    lines = sht.max_row
    cols = sht.max_column
    print("行数",lines)
    print("列数",cols)

    list = ["key1", "key2", "key3"]
    sum = lines - 1
    count = 0

    for i in range(1, lines + 1):
        row = sht[i]
        hasKeyFlag = False
        for j in range(1, cols + 1):
            if(not hasKeyFlag):
                for m in list:
                    if (row[j - 1].value and m in str(row[j - 1].value)):
                        for k in range(1, cols + 1):
                            row[k - 1].fill = fill_1
                        count = count + 1
                        hasKeyFlag = True
                        break
            else:
                break
    print("包含关键词",list,"的记录数为",count,"总记录数为",sum,"占比：",round(count / sum,4) * 100,'%')

    wb.save(excel_address)

if __name__ == '__main__':
    handelExcel()