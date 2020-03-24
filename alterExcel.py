from openpyxl import load_workbook
from openpyxl.styles import PatternFill, colors

excel_address = r"cases.xlsx"
wb = load_workbook(excel_address)
sht = wb.worksheets[0]

sht["A1"] = "测试"
sht["A3"] = "测试"
sht["A5"] = "测试"
sht["A7"] = "测试"

fill_1 = PatternFill("solid", fgColor="1874CD")
fill_2 = PatternFill("solid", fgColor="BCEE68")
fill_3 = PatternFill("solid", fgColor=colors.RED)
fill_4 = PatternFill("solid", fgColor=colors.GREEN)

sht["A1"].fill = fill_1
sht["A3"].fill = fill_2
sht["A5"].fill = fill_3
sht["A7"].fill = fill_4

row = sht[1]
# row.fill = fill_1
row[1].fill = fill_1
print(row[1].value)

wb.save(excel_address)
